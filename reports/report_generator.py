"""
Módulo de Generación de Reportes
Crea reportes en Excel con gráficos y estadísticas
"""

from datetime import datetime, timedelta
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter


class ReportGenerator:
    """Generador de reportes en Excel"""
    
    def __init__(self, history_manager):
        """
        Inicializar generador de reportes
        
        Args:
            history_manager: Instancia de InstallationHistory
        """
        self.history = history_manager
    
    def generate_monthly_report(self, year, month, output_path=None):
        """
        Generar reporte mensual
        
        Args:
            year: Año
            month: Mes (1-12)
            output_path: Ruta de salida (opcional)
            
        Returns:
            Ruta del archivo generado
        """
        # Calcular fechas del mes
        start_date = datetime(year, month, 1).isoformat()
        
        if month == 12:
            end_date = datetime(year + 1, 1, 1).isoformat()
        else:
            end_date = datetime(year, month + 1, 1).isoformat()
        
        # Nombre del mes en español
        month_names = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                      'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
        month_name = month_names[month - 1]
        
        # Ruta de salida
        if output_path is None:
            output_path = Path.home() / "Downloads" / f"Reporte_{month_name}_{year}.xlsx"
        
        # Obtener datos
        installations = self.history.get_installations(start_date=start_date, end_date=end_date)
        stats = self.history.get_statistics(start_date=start_date, end_date=end_date)
        
        if installations is None:
            installations = []
            
        if stats is None:
            stats = {
                'total_installations': 0,
                'successful_installations': 0,
                'failed_installations': 0,
                'success_rate': 0,
                'average_time_minutes': 0,
                'unique_clients': 0,
                'top_drivers': {},
                'by_brand': {}
            }
        
        # Crear workbook
        wb = openpyxl.Workbook()
        
        # Hoja 1: Resumen
        self._create_summary_sheet(wb, installations, stats, f"{month_name} {year}")
        
        # Hoja 2: Detalle de Instalaciones
        self._create_installations_sheet(wb, installations)
        
        # Hoja 3: Por Cliente
        self._create_clients_sheet(wb, installations)
        
        # Hoja 4: Gráficos
        self._create_charts_sheet(wb, stats)
        
        # Guardar
        wb.save(output_path)
        
        return str(output_path)

    def generate_yearly_report(self, year, output_path=None):
        """
        Generar reporte anual.

        Args:
            year: Año
            output_path: Ruta de salida (opcional)

        Returns:
            Ruta del archivo generado
        """
        start_date = datetime(year, 1, 1).isoformat()
        end_date = datetime(year + 1, 1, 1).isoformat()

        if output_path is None:
            output_path = Path.home() / "Downloads" / f"Reporte_Anual_{year}.xlsx"

        installations = self.history.get_installations(start_date=start_date, end_date=end_date)
        stats = self.history.get_statistics(start_date=start_date, end_date=end_date)

        if installations is None:
            installations = []

        if stats is None:
            stats = {
                'total_installations': 0,
                'successful_installations': 0,
                'failed_installations': 0,
                'success_rate': 0,
                'average_time_minutes': 0,
                'unique_clients': 0,
                'top_drivers': {},
                'by_brand': {}
            }

        wb = openpyxl.Workbook()
        self._create_summary_sheet(wb, installations, stats, f"Anual {year}")
        self._create_installations_sheet(wb, installations)
        self._create_clients_sheet(wb, installations)
        self._create_charts_sheet(wb, stats)

        wb.save(output_path)
        return str(output_path)
    
    def generate_daily_report(self, date=None, output_path=None):
        """
        Generar reporte del día
        
        Args:
            date: Fecha (datetime object, default=hoy)
            output_path: Ruta de salida
            
        Returns:
            Ruta del archivo generado
        """
        if date is None:
            date = datetime.now()
        
        start_date = date.replace(hour=0, minute=0, second=0, microsecond=0).isoformat()
        end_date = date.replace(hour=23, minute=59, second=59, microsecond=999999).isoformat()
        
        if output_path is None:
            date_str = date.strftime("%Y-%m-%d")
            output_path = Path.home() / "Downloads" / f"Reporte_Diario_{date_str}.xlsx"
        
        installations = self.history.get_installations(start_date=start_date, end_date=end_date)
        stats = self.history.get_statistics(start_date=start_date, end_date=end_date)
        
        if installations is None:
            installations = []
            
        if stats is None:
            stats = {
                'total_installations': 0,
                'successful_installations': 0,
                'failed_installations': 0,
                'success_rate': 0,
                'average_time_minutes': 0,
                'unique_clients': 0,
                'top_drivers': {},
                'by_brand': {}
            }
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte Diario"
        
        # Título
        title = f"Reporte Diario - {date.strftime('%d/%m/%Y')}"
        self._add_title(ws, title)
        
        # Resumen rápido
        row = 3
        ws[f'A{row}'] = "Total Instalaciones:"
        ws[f'B{row}'] = stats['total_installations']
        ws[f'B{row}'].font = Font(bold=True, size=14)
        
        row += 1
        ws[f'A{row}'] = "Exitosas:"
        ws[f'B{row}'] = stats['successful_installations']
        ws[f'B{row}'].font = Font(color="008000")
        
        row += 1
        ws[f'A{row}'] = "Fallidas:"
        ws[f'B{row}'] = stats['failed_installations']
        ws[f'B{row}'].font = Font(color="FF0000")
        
        row += 2
        
        # Detalle de instalaciones
        if installations:
            headers = ['Hora', 'Cliente', 'Marca', 'Versión', 'Estado', 'Tiempo (min)', 'Notas']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row, col, header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
            
            row += 1
            
            for inst in installations:
                timestamp = datetime.fromisoformat(inst['timestamp'])
                time_str = timestamp.strftime('%H:%M')
                
                time_minutes = inst['installation_time_seconds'] / 60 if inst['installation_time_seconds'] else 0
                
                ws.cell(row, 1, time_str)
                ws.cell(row, 2, inst.get('client_name') or 'N/A')
                ws.cell(row, 3, inst['driver_brand'])
                ws.cell(row, 4, inst['driver_version'])
                ws.cell(row, 5, '✓' if inst['status'] == 'success' else '✗')
                ws.cell(row, 6, round(time_minutes, 1))
                ws.cell(row, 7, inst.get('notes') or '')
                
                # Colorear estado
                status_cell = ws.cell(row, 5)
                if inst['status'] == 'success':
                    status_cell.font = Font(color="008000", bold=True)
                else:
                    status_cell.font = Font(color="FF0000", bold=True)
                
                row += 1
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 40
        
        wb.save(output_path)
        return str(output_path)
    
    def generate_client_report(self, client_name, output_path=None):
        """
        Generar reporte de un cliente específico
        
        Args:
            client_name: Nombre del cliente
            output_path: Ruta de salida
            
        Returns:
            Ruta del archivo generado
        """
        if output_path is None:
            safe_name = "".join(c for c in client_name if c.isalnum() or c in (' ', '_')).strip()
            output_path = Path.home() / "Downloads" / f"Reporte_Cliente_{safe_name}.xlsx"
        
        history = self.history.get_client_history(client_name)
        
        if history is None:
            history = {
                'client': None,
                'installations': [],
                'notes': []
            }
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Historial Cliente"
        
        # Título
        self._add_title(ws, f"Historial de {client_name}")
        
        # Información del cliente
        row = 3
        if history['client']:
            client = history['client']
            ws[f'A{row}'] = "Total Servicios:"
            ws[f'B{row}'] = client.get('total_services', 0)
            
            row += 1
            ws[f'A{row}'] = "Última Visita:"
            if client.get('last_visit'):
                last_visit = datetime.fromisoformat(client['last_visit'])
                ws[f'B{row}'] = last_visit.strftime('%d/%m/%Y %H:%M')
            
            row += 1
            ws[f'A{row}'] = "Contacto:"
            ws[f'B{row}'] = client.get('contact') or 'N/A'
            
            row += 1
            ws[f'A{row}'] = "Dirección:"
            ws[f'B{row}'] = client.get('address') or 'N/A'
        
        row += 2
        
        # Historial de instalaciones
        ws[f'A{row}'] = "Historial de Instalaciones"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        if history['installations']:
            headers = ['Fecha', 'Marca', 'Versión', 'Estado', 'Tiempo', 'Notas']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row, col, header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
            
            row += 1
            
            for inst in history['installations']:
                timestamp = datetime.fromisoformat(inst['timestamp'])
                date_str = timestamp.strftime('%d/%m/%Y %H:%M')
                
                time_str = ''
                if inst['installation_time_seconds']:
                    minutes = inst['installation_time_seconds'] / 60
                    time_str = f"{minutes:.1f} min"
                
                ws.cell(row, 1, date_str)
                ws.cell(row, 2, inst['driver_brand'])
                ws.cell(row, 3, inst['driver_version'])
                ws.cell(row, 4, '✓ Exitosa' if inst['status'] == 'success' else '✗ Fallida')
                ws.cell(row, 5, time_str)
                ws.cell(row, 6, inst.get('notes') or '')
                
                row += 1
        
        row += 2
        
        # Notas del cliente
        if history['notes']:
            ws[f'A{row}'] = "Notas y Observaciones"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            headers = ['Fecha', 'Categoría', 'Nota']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row, col, header)
                cell.font = Font(bold=True)
            
            row += 1
            
            for note in history['notes']:
                timestamp = datetime.fromisoformat(note['timestamp'])
                date_str = timestamp.strftime('%d/%m/%Y %H:%M')
                
                ws.cell(row, 1, date_str)
                ws.cell(row, 2, note.get('category') or 'General')
                ws.cell(row, 3, note.get('note') or '')
                
                row += 1
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 50
        
        wb.save(output_path)
        return str(output_path)
    
    def _create_summary_sheet(self, wb, installations, stats, period_name):
        """Crear hoja de resumen"""
        ws = wb.active
        ws.title = "Resumen"
        
        # Título
        self._add_title(ws, f"Reporte de Instalaciones - {period_name}")
        
        # Estadísticas principales
        row = 3
        data = [
            ('Total de Instalaciones', stats['total_installations']),
            ('Instalaciones Exitosas', stats['successful_installations']),
            ('Instalaciones Fallidas', stats['failed_installations']),
            ('Tasa de Éxito', f"{stats['success_rate']}%"),
            ('Tiempo Promedio', f"{stats['average_time_minutes']:.1f} minutos"),
            ('Clientes Únicos', stats['unique_clients']),
        ]
        
        for label, value in data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            ws[f'B{row}'].font = Font(size=12)
            row += 1
        
        row += 2
        
        # Top drivers
        ws[f'A{row}'] = "Drivers Más Instalados"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        headers = ['Driver', 'Cantidad']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            self._style_header(cell)
        
        row += 1
        
        for driver, count in stats['top_drivers'].items():
            ws.cell(row, 1, driver)
            ws.cell(row, 2, count)
            row += 1
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
    
    def _create_installations_sheet(self, wb, installations):
        """Crear hoja de detalle de instalaciones"""
        ws = wb.create_sheet("Instalaciones")
        
        # Encabezados
        headers = ['Fecha/Hora', 'Cliente', 'PC', 'Marca', 'Versión', 
                  'Estado', 'Tiempo (min)', 'Técnico', 'Notas']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            self._style_header(cell)
        
        # Datos
        row = 2
        for inst in installations:
            timestamp = datetime.fromisoformat(inst['timestamp'])
            date_str = timestamp.strftime('%d/%m/%Y %H:%M')
            
            time_minutes = inst['installation_time_seconds'] / 60 if inst['installation_time_seconds'] else 0
            
            ws.cell(row, 1, date_str)
            ws.cell(row, 2, inst.get('client_name') or 'N/A')
            ws.cell(row, 3, inst.get('client_pc_name') or 'N/A')
            ws.cell(row, 4, inst['driver_brand'])
            ws.cell(row, 5, inst['driver_version'])
            ws.cell(row, 6, 'Exitosa' if inst['status'] == 'success' else 'Fallida')
            ws.cell(row, 7, round(time_minutes, 1))
            ws.cell(row, 8, inst.get('technician_name') or 'N/A')
            ws.cell(row, 9, inst.get('notes') or '')
            
            # Colorear estado
            status_cell = ws.cell(row, 6)
            if inst['status'] == 'success':
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            else:
                status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            row += 1
        
        # Ajustar anchos
        widths = [18, 25, 20, 15, 12, 12, 12, 20, 40]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
    
    def _create_clients_sheet(self, wb, installations):
        """Crear hoja de resumen por cliente"""
        ws = wb.create_sheet("Por Cliente")
        
        # Agrupar por cliente
        clients_data = {}
        for inst in installations:
            client = inst.get('client_name') or 'Sin nombre'
            if client not in clients_data:
                clients_data[client] = {
                    'total': 0,
                    'successful': 0,
                    'failed': 0,
                    'drivers': []
                }
            
            clients_data[client]['total'] += 1
            if inst['status'] == 'success':
                clients_data[client]['successful'] += 1
            else:
                clients_data[client]['failed'] += 1
            
            driver_name = f"{inst['driver_brand']} {inst['driver_version']}"
            clients_data[client]['drivers'].append(driver_name)
        
        # Encabezados
        headers = ['Cliente', 'Total', 'Exitosas', 'Fallidas', 'Tasa Éxito', 'Drivers']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            self._style_header(cell)
        
        # Datos
        row = 2
        for client, data in sorted(clients_data.items(), key=lambda x: x[1]['total'], reverse=True):
            success_rate = (data['successful'] / data['total'] * 100) if data['total'] > 0 else 0
            drivers_list = ', '.join(set(data['drivers']))
            
            ws.cell(row, 1, client)
            ws.cell(row, 2, data['total'])
            ws.cell(row, 3, data['successful'])
            ws.cell(row, 4, data['failed'])
            ws.cell(row, 5, f"{success_rate:.1f}%")
            ws.cell(row, 6, drivers_list)
            
            row += 1
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 50
    
    def _create_charts_sheet(self, wb, stats):
        """Crear hoja con gráficos"""
        ws = wb.create_sheet("Gráficos")
        
        # Datos para gráfico de marcas
        row = 2
        ws.cell(row, 1, "Marca")
        ws.cell(row, 2, "Cantidad")
        self._style_header(ws.cell(row, 1))
        self._style_header(ws.cell(row, 2))
        
        row += 1
        start_row = row
        
        for brand, count in stats['by_brand'].items():
            ws.cell(row, 1, brand)
            ws.cell(row, 2, count)
            row += 1
        
        end_row = row - 1
        
        # Crear gráfico de barras
        if end_row >= start_row:
            chart = BarChart()
            chart.title = "Instalaciones por Marca"
            chart.x_axis.title = "Marca"
            chart.y_axis.title = "Cantidad"
            
            data = Reference(ws, min_col=2, min_row=2, max_row=end_row)
            cats = Reference(ws, min_col=1, min_row=start_row, max_row=end_row)
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            
            ws.add_chart(chart, "D2")
    
    def _add_title(self, ws, title):
        """Agregar título a la hoja"""
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')
        ws['A1'].alignment = Alignment(horizontal='center')
    
    def _style_header(self, cell):
        """Aplicar estilo a encabezado"""
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
