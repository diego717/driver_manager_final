import tempfile
import unittest
from datetime import datetime
from pathlib import Path
import uuid
from unittest.mock import MagicMock

import openpyxl

from reports.report_generator import ReportGenerator


class TestReportGenerator(unittest.TestCase):
    def _temp_xlsx_path(self, prefix):
        return Path(tempfile.gettempdir()) / f"{prefix}_{uuid.uuid4().hex}.xlsx"

    def _cleanup_file(self, path):
        try:
            if path.exists():
                path.unlink()
        except PermissionError:
            # On Windows, openpyxl may release the handle slightly later.
            pass

    def _sample_installations(self):
        return [
            {
                "timestamp": "2026-02-13T10:00:00",
                "client_name": "Cliente A",
                "client_pc_name": "PC-01",
                "driver_brand": "Zebra",
                "driver_version": "1.2.3",
                "status": "success",
                "installation_time_seconds": 120,
                "technician_name": "Diego",
                "notes": "OK",
            },
            {
                "timestamp": "2026-02-13T11:30:00",
                "client_name": "Cliente B",
                "client_pc_name": "PC-02",
                "driver_brand": "Magicard",
                "driver_version": "2.0.0",
                "status": "failed",
                "installation_time_seconds": 0,
                "technician_name": "Ana",
                "notes": "Timeout",
            },
        ]

    def _sample_stats(self):
        return {
            "total_installations": 2,
            "successful_installations": 1,
            "failed_installations": 1,
            "success_rate": 50,
            "average_time_minutes": 1.0,
            "unique_clients": 2,
            "top_drivers": {"Zebra 1.2.3": 1, "Magicard 2.0.0": 1},
            "by_brand": {"Zebra": 1, "Magicard": 1},
        }

    def test_generate_daily_report_creates_file_with_expected_content(self):
        history = MagicMock()
        history.get_installations.return_value = self._sample_installations()
        history.get_statistics.return_value = self._sample_stats()
        generator = ReportGenerator(history)

        output = self._temp_xlsx_path("daily")
        try:
            result = generator.generate_daily_report(
                date=datetime(2026, 2, 13),
                output_path=output,
            )

            self.assertEqual(result, str(output))
            self.assertTrue(output.exists())

            wb = openpyxl.load_workbook(output, read_only=True, data_only=True)
            try:
                ws = wb["Reporte Diario"]
                self.assertIn("Reporte Diario", ws["A1"].value)
                self.assertEqual(ws["B3"].value, 2)
                self.assertEqual(ws["C8"].value, "Zebra")
                self.assertEqual(ws["F8"].value, 2)
            finally:
                wb.close()
        finally:
            self._cleanup_file(output)

    def test_generate_monthly_report_uses_fallback_when_history_returns_none(self):
        history = MagicMock()
        history.get_installations.return_value = None
        history.get_statistics.return_value = None
        generator = ReportGenerator(history)

        output = self._temp_xlsx_path("monthly")
        try:
            result = generator.generate_monthly_report(2026, 2, output_path=output)

            self.assertEqual(result, str(output))
            self.assertTrue(output.exists())

            wb = openpyxl.load_workbook(output, read_only=True, data_only=True)
            try:
                self.assertIn("Resumen", wb.sheetnames)
                self.assertIn("Instalaciones", wb.sheetnames)
                self.assertIn("Por Cliente", wb.sheetnames)
                self.assertEqual(len(wb.sheetnames), 4)
                ws = wb["Resumen"]
                self.assertEqual(ws["B3"].value, 0)
            finally:
                wb.close()
        finally:
            self._cleanup_file(output)

    def test_generate_client_report_writes_client_history(self):
        history = MagicMock()
        history.get_client_history.return_value = {
            "client": {
                "total_services": 3,
                "last_visit": "2026-02-13T10:30:00",
                "contact": "John",
                "address": "Street 123",
            },
            "installations": [
                {
                    "timestamp": "2026-02-13T10:00:00",
                    "driver_brand": "Zebra",
                    "driver_version": "1.2.3",
                    "status": "success",
                    "installation_time_seconds": 180,
                    "notes": "OK",
                }
            ],
            "notes": [
                {
                    "timestamp": "2026-02-13T12:00:00",
                    "category": "General",
                    "note": "Todo bien",
                }
            ],
        }
        generator = ReportGenerator(history)

        output = self._temp_xlsx_path("client")
        try:
            result = generator.generate_client_report("Cliente A", output_path=output)

            self.assertEqual(result, str(output))
            self.assertTrue(output.exists())

            wb = openpyxl.load_workbook(output, read_only=True, data_only=True)
            try:
                ws = wb["Historial Cliente"]
                self.assertIn("Historial de Cliente A", ws["A1"].value)

                rows = list(ws.iter_rows(min_row=1, max_col=6, values_only=True))
                self.assertTrue(any(row[1] == "Zebra" for row in rows if row))
                self.assertTrue(
                    any(
                        isinstance(row[3], str) and "Exitosa" in row[3]
                        for row in rows
                        if row and len(row) > 3
                    )
                )
            finally:
                wb.close()
        finally:
            self._cleanup_file(output)

    def test_generate_yearly_report_creates_output_file(self):
        history = MagicMock()
        history.get_installations.return_value = self._sample_installations()
        history.get_statistics.return_value = self._sample_stats()
        generator = ReportGenerator(history)

        output = self._temp_xlsx_path("yearly")
        try:
            result = generator.generate_yearly_report(2026, output_path=output)

            self.assertEqual(result, str(output))
            self.assertTrue(output.exists())

            wb = openpyxl.load_workbook(output, read_only=True, data_only=True)
            try:
                ws = wb["Resumen"]
                self.assertIn("Anual 2026", ws["A1"].value)
                self.assertEqual(ws["B3"].value, 2)
            finally:
                wb.close()
        finally:
            self._cleanup_file(output)


if __name__ == "__main__":
    unittest.main()
